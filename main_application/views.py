from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Avg, Q, Sum
from django.utils import timezone
from datetime import timedelta
from .models import (
    User, Candidate, ExamResult, AggregateResult, School, 
    AcademicYear, Subject, ResultAccessPayment, FraudAttemptLog,
    UserActivityLog, SchoolPerformanceReport, EducationLevel
)
import json


def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def log_activity(request, user, action, description):
    """Log user activity"""
    UserActivityLog.objects.create(
        user=user,
        action=action,
        description=description,
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
    )


def login_view(request):
    """
    Unified login view for all user types
    Redirects to appropriate dashboard based on user role
    """
    # Redirect if already logged in
    if request.user.is_authenticated:
        return redirect_to_dashboard(request.user)
    
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        remember_me = request.POST.get('remember_me', False)
        
        # Authenticate user
        user = authenticate(request, email=email, password=password)
        
        if user is not None:
            # Check if account is active
            if not user.is_active:
                messages.error(request, 'Your account has been deactivated. Please contact support.')
                return render(request, 'main_application/login.html')
            
            # Check if account has expired (for marks entry clerks)
            if user.is_account_expired:
                messages.error(request, 'Your account has expired. Please contact KNEC staff for renewal.')
                return render(request, 'main_application/login.html')
            
            # Check account expiry
            if user.account_expires_at and timezone.now() > user.account_expires_at:
                user.is_account_expired = True
                user.is_active = False
                user.save()
                messages.error(request, 'Your account has expired. Please contact KNEC staff for renewal.')
                return render(request, 'main_application/login.html')
            
            # Login user
            login(request, user)
            
            # Set session expiry
            if not remember_me:
                request.session.set_expiry(0)  # Session expires when browser closes
            else:
                request.session.set_expiry(1209600)  # 2 weeks
            
            # Update last login IP
            user.last_login_ip = get_client_ip(request)
            user.save(update_fields=['last_login_ip'])
            
            # Log activity
            log_activity(request, user, 'LOGIN', f'User {user.email} logged in successfully')
            
            # Success message
            messages.success(request, f'Welcome back, {user.get_full_name()}!')
            
            # Redirect to appropriate dashboard
            return redirect_to_dashboard(user)
        else:
            # Log failed attempt
            ip_address = get_client_ip(request)
            FraudAttemptLog.objects.create(
                attempt_type='UNAUTHORIZED_ACCESS',
                ip_address=ip_address,
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
                description=f'Failed login attempt for email: {email}'
            )
            messages.error(request, 'Invalid email or password. Please try again.')
    
    return render(request, 'auth/login.html')


def redirect_to_dashboard(user):
    """Redirect user to appropriate dashboard based on role"""
    dashboard_mapping = {
        'ADMIN': 'admin_dashboard',
        'KNEC_STAFF': 'knec_dashboard',
        'MARKS_ENTRY': 'marks_entry_dashboard',
        'SCHOOL_ADMIN': 'school_admin_dashboard',
        'SCHOOL_STAFF': 'school_staff_dashboard',
    }
    
    dashboard_url = dashboard_mapping.get(user.user_type, 'admin_dashboard')
    return redirect(dashboard_url)


@login_required
def logout_view(request):
    """Logout view"""
    user = request.user
    log_activity(request, user, 'LOGOUT', f'User {user.email} logged out')
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


# ============================================================
# ADMIN DASHBOARD
# ============================================================

@login_required
def admin_dashboard(request):
    """
    Admin Dashboard with comprehensive statistics and graphs
    """
    # Check if user is admin
    if not request.user.is_admin:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect_to_dashboard(request.user)
    
    # Get active academic year
    active_year = AcademicYear.objects.filter(is_active=True).first()
    
    # Get date ranges
    today = timezone.now()
    last_30_days = today - timedelta(days=30)
    last_7_days = today - timedelta(days=7)
    
    # ===== STATISTICS =====
    
    # User Statistics
    total_users = User.objects.filter(is_active=True).count()
    admin_users = User.objects.filter(user_type='ADMIN', is_active=True).count()
    knec_staff = User.objects.filter(user_type='KNEC_STAFF', is_active=True).count()
    marks_entry_clerks = User.objects.filter(user_type='MARKS_ENTRY', is_active=True).count()
    school_users = User.objects.filter(
        user_type__in=['SCHOOL_ADMIN', 'SCHOOL_STAFF'], 
        is_active=True
    ).count()
    
    # School Statistics
    total_schools = School.objects.filter(is_active=True).count()
    primary_schools = School.objects.filter(
        category__name='PRIMARY', 
        is_active=True
    ).count()
    secondary_schools = School.objects.filter(
        category__name__in=['JSS', 'SENIOR'], 
        is_active=True
    ).count()
    
    # Candidate Statistics
    total_candidates = Candidate.objects.filter(is_active=True).count()
    if active_year:
        current_year_candidates = Candidate.objects.filter(
            academic_year=active_year,
            is_active=True
        ).count()
    else:
        current_year_candidates = 0
    
    # Results Statistics
    total_results = ExamResult.objects.count()
    released_results = AggregateResult.objects.filter(is_released=True).count()
    pending_results = AggregateResult.objects.filter(is_released=False).count()
    
    # Payment Statistics
    total_payments = ResultAccessPayment.objects.filter(status='completed').count()
    total_revenue = ResultAccessPayment.objects.filter(
        status='completed'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    recent_payments = ResultAccessPayment.objects.filter(
        status='completed',
        payment_date__gte=last_30_days
    ).count()
    
    # Fraud Statistics
    total_fraud_attempts = FraudAttemptLog.objects.count()
    unresolved_fraud = FraudAttemptLog.objects.filter(is_resolved=False).count()
    recent_fraud = FraudAttemptLog.objects.filter(
        created_at__gte=last_7_days
    ).count()
    
    # ===== GRAPH DATA =====
    
    # 1. Candidates Registration Trend (Last 6 months)
    six_months_ago = today - timedelta(days=180)
    candidates_by_month = []
    for i in range(6):
        month_start = today - timedelta(days=30 * (5 - i))
        month_end = today - timedelta(days=30 * (4 - i))
        count = Candidate.objects.filter(
            registration_date__gte=month_start,
            registration_date__lt=month_end
        ).count()
        candidates_by_month.append({
            'month': month_start.strftime('%b'),
            'count': count
        })
    
    # 2. Results by Education Level
    results_by_level = []
    for level in EducationLevel.objects.filter(is_active=True):
        count = AggregateResult.objects.filter(
            candidate__education_level=level,
            is_released=True
        ).count()
        results_by_level.append({
            'level': level.get_name_display(),
            'count': count
        })
    
    # 3. Grade Distribution (Overall)
    grade_distribution = AggregateResult.objects.filter(
        is_released=True
    ).values('mean_grade').annotate(
        count=Count('id')
    ).order_by('mean_grade')
    
    grades_data = []
    grade_order = ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'E']
    for grade in grade_order:
        count = next((item['count'] for item in grade_distribution if item['mean_grade'] == grade), 0)
        grades_data.append({
            'grade': grade,
            'count': count
        })
    
    # 4. School Performance (Top 10 Schools)
    top_schools = []
    if active_year:
        school_reports = SchoolPerformanceReport.objects.filter(
            academic_year=active_year
        ).select_related('school').order_by('-mean_score')[:10]
        
        for report in school_reports:
            top_schools.append({
                'school': report.school.name,
                'mean_score': float(report.mean_score)
            })
    
    # 5. User Activity (Last 7 days)
    activity_by_day = []
    for i in range(7):
        day = today - timedelta(days=6 - i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        count = UserActivityLog.objects.filter(
            timestamp__gte=day_start,
            timestamp__lte=day_end
        ).count()
        
        activity_by_day.append({
            'day': day.strftime('%a'),
            'count': count
        })
    
    # 6. Payment Trends (Last 30 days)
    payment_by_day = []
    for i in range(30, 0, -3):  # Every 3 days
        day = today - timedelta(days=i)
        day_end = today - timedelta(days=i - 3)
        
        amount = ResultAccessPayment.objects.filter(
            status='completed',
            payment_date__gte=day,
            payment_date__lt=day_end
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        payment_by_day.append({
            'date': day.strftime('%d %b'),
            'amount': float(amount)
        })
    
    # 7. User Distribution by Type
    user_type_distribution = [
        {'type': 'Admin', 'count': admin_users},
        {'type': 'KNEC Staff', 'count': knec_staff},
        {'type': 'Marks Entry', 'count': marks_entry_clerks},
        {'type': 'School Users', 'count': school_users},
    ]
    
    # 8. Recent Activities
    recent_activities = UserActivityLog.objects.select_related('user').order_by('-timestamp')[:10]
    
    # 9. Recent Registrations
    recent_candidates = Candidate.objects.select_related(
        'school', 'education_level', 'academic_year'
    ).order_by('-registration_date')[:10]
    
    # 10. Pending Tasks
    pending_tasks = {
        'unverified_certificates': Candidate.objects.filter(
            is_birth_cert_verified=False,
            birth_certificate__isnull=False
        ).count(),
        'pending_results': pending_results,
        'unresolved_fraud': unresolved_fraud,
        'expired_accounts': User.objects.filter(
            is_account_expired=True,
            is_active=False
        ).count(),
    }
    
    context = {
        # Statistics
        'total_users': total_users,
        'total_schools': total_schools,
        'total_candidates': total_candidates,
        'current_year_candidates': current_year_candidates,
        'total_results': total_results,
        'released_results': released_results,
        'pending_results': pending_results,
        'total_payments': total_payments,
        'total_revenue': total_revenue,
        'recent_payments': recent_payments,
        'total_fraud_attempts': total_fraud_attempts,
        'unresolved_fraud': unresolved_fraud,
        'recent_fraud': recent_fraud,
        
        # Breakdown
        'admin_users': admin_users,
        'knec_staff': knec_staff,
        'marks_entry_clerks': marks_entry_clerks,
        'school_users': school_users,
        'primary_schools': primary_schools,
        'secondary_schools': secondary_schools,
        
        # Graph Data (JSON)
        'candidates_by_month_json': json.dumps(candidates_by_month),
        'results_by_level_json': json.dumps(results_by_level),
        'grades_data_json': json.dumps(grades_data),
        'top_schools_json': json.dumps(top_schools),
        'activity_by_day_json': json.dumps(activity_by_day),
        'payment_by_day_json': json.dumps(payment_by_day),
        'user_type_distribution_json': json.dumps(user_type_distribution),
        
        # Lists
        'recent_activities': recent_activities,
        'recent_candidates': recent_candidates,
        'pending_tasks': pending_tasks,
        
        # Other
        'active_year': active_year,
    }
    
    return render(request, 'dashboards/admin_dashboard.html', context)


# ============================================================
# OTHER DASHBOARDS (Placeholders - Implement as needed)
# ============================================================

@login_required
def knec_dashboard(request):
    """KNEC Staff Dashboard"""
    if not request.user.is_knec_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect_to_dashboard(request.user)
    
    context = {
        'page_title': 'KNEC Staff Dashboard',
    }
    return render(request, 'dashboards/knec_dashboard.html', context)


@login_required
def marks_entry_dashboard(request):
    """Marks Entry Clerk Dashboard"""
    if not request.user.is_marks_entry:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect_to_dashboard(request.user)
    
    context = {
        'page_title': 'Marks Entry Dashboard',
    }
    return render(request, 'dashboards/marks_entry_dashboard.html', context)


@login_required
def school_admin_dashboard(request):
    """School Administrator Dashboard"""
    if not request.user.is_school_user:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect_to_dashboard(request.user)
    
    context = {
        'page_title': 'School Administrator Dashboard',
    }
    return render(request, 'dashboards/school_admin_dashboard.html', context)


@login_required
def school_staff_dashboard(request):
    """School Staff Dashboard"""
    if not request.user.is_school_user:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect_to_dashboard(request.user)
    
    context = {
        'page_title': 'School Staff Dashboard',
    }
    return render(request, 'dashboards/school_staff_dashboard.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    Candidate, School, EducationLevel, AcademicYear, 
    BirthCertificateRegistry, User, UserActivityLog
)


def log_activity(user, action, description, ip_address, user_agent=''):
    """Helper function to log user activities"""
    UserActivityLog.objects.create(
        user=user,
        action=action,
        description=description,
        ip_address=ip_address,
        user_agent=user_agent
    )


def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


@login_required
def candidates_list(request):
    """List all candidates with search, filter, and pagination"""
    
    # Base queryset
    candidates = Candidate.objects.select_related(
        'school', 'education_level', 'academic_year', 'birth_certificate', 'registered_by'
    ).all()
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    school_filter = request.GET.get('school', '')
    level_filter = request.GET.get('level', '')
    year_filter = request.GET.get('year', '')
    gender_filter = request.GET.get('gender', '')
    status_filter = request.GET.get('status', '')
    cert_verified = request.GET.get('cert_verified', '')
    
    # Apply search
    if search_query:
        candidates = candidates.filter(
            Q(index_number__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(middle_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(birth_certificate__certificate_number__icontains=search_query) |
            Q(school__name__icontains=search_query) |
            Q(school__code__icontains=search_query)
        )
    
    # Apply filters
    if school_filter:
        candidates = candidates.filter(school_id=school_filter)
    
    if level_filter:
        candidates = candidates.filter(education_level_id=level_filter)
    
    if year_filter:
        candidates = candidates.filter(academic_year_id=year_filter)
    
    if gender_filter:
        candidates = candidates.filter(gender=gender_filter)
    
    if status_filter:
        is_active = status_filter == 'active'
        candidates = candidates.filter(is_active=is_active)
    
    if cert_verified:
        is_verified = cert_verified == 'verified'
        candidates = candidates.filter(is_birth_cert_verified=is_verified)
    
    # Get filter options
    schools = School.objects.filter(is_active=True).order_by('name')
    education_levels = EducationLevel.objects.filter(is_active=True)
    academic_years = AcademicYear.objects.all().order_by('-year')
    
    # Statistics
    total_candidates = candidates.count()
    active_candidates = candidates.filter(is_active=True).count()
    verified_certs = candidates.filter(is_birth_cert_verified=True).count()
    
    # Export to Excel
    if request.GET.get('export') == 'excel':
        return export_candidates_to_excel(candidates, request)
    
    # Pagination
    paginator = Paginator(candidates.order_by('-registration_date'), 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Log activity
    log_activity(
        user=request.user,
        action='VIEW',
        description=f'Viewed candidates list (Total: {total_candidates})',
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    context = {
        'page_obj': page_obj,
        'schools': schools,
        'education_levels': education_levels,
        'academic_years': academic_years,
        'total_candidates': total_candidates,
        'active_candidates': active_candidates,
        'verified_certs': verified_certs,
        'search_query': search_query,
        'filters': {
            'school': school_filter,
            'level': level_filter,
            'year': year_filter,
            'gender': gender_filter,
            'status': status_filter,
            'cert_verified': cert_verified,
        },
        'now': timezone.now(),
    }
    
    return render(request, 'candidates/candidates_list.html', context)


@login_required
def candidate_detail(request, index_number):
    """View candidate details"""
    candidate = get_object_or_404(
        Candidate.objects.select_related(
            'school', 'education_level', 'academic_year', 
            'birth_certificate', 'registered_by'
        ),
        index_number=index_number
    )
    
    # Get related data
    exam_results = candidate.exam_results.select_related(
        'subject', 'grading_scheme_used', 'entered_by'
    ).all()
    
    aggregate_result = None
    try:
        aggregate_result = candidate.aggregate_result
    except:
        pass
    
    payments = candidate.result_payments.all().order_by('-created_at')
    
    # Log activity
    log_activity(
        user=request.user,
        action='VIEW',
        description=f'Viewed candidate details: {candidate.index_number}',
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    context = {
        'candidate': candidate,
        'exam_results': exam_results,
        'aggregate_result': aggregate_result,
        'payments': payments,
        'now': timezone.now(),
    }
    
    return render(request, 'candidates/candidate_detail.html', context)


@login_required
def candidate_create(request):
    """Create new candidate"""
    
    if request.method == 'POST':
        try:
            # Get form data
            school_id = request.POST.get('school')
            education_level_id = request.POST.get('education_level')
            academic_year_id = request.POST.get('academic_year')
            first_name = request.POST.get('first_name')
            middle_name = request.POST.get('middle_name', '')
            last_name = request.POST.get('last_name')
            gender = request.POST.get('gender')
            date_of_birth = request.POST.get('date_of_birth')
            phone_number = request.POST.get('phone_number', '')
            parent_guardian_phone = request.POST.get('parent_guardian_phone')
            birth_cert_number = request.POST.get('birth_certificate_number', '')
            
            # Validate required fields
            if not all([school_id, education_level_id, academic_year_id, 
                       first_name, last_name, gender, date_of_birth, parent_guardian_phone]):
                messages.error(request, 'Please fill in all required fields.')
                return redirect('candidate_create')
            
            # Get related objects
            school = School.objects.get(id=school_id)
            education_level = EducationLevel.objects.get(id=education_level_id)
            academic_year = AcademicYear.objects.get(id=academic_year_id)
            
            # Check age for birth certificate requirement
            dob = datetime.strptime(date_of_birth, '%Y-%m-%d').date()
            age = (timezone.now().date() - dob).days / 365.25
            
            birth_certificate = None
            is_birth_cert_verified = False
            
            if age < 18 and birth_cert_number:
                # Verify birth certificate
                try:
                    birth_certificate = BirthCertificateRegistry.objects.get(
                        certificate_number=birth_cert_number
                    )
                    
                    # Check if already used
                    if birth_certificate.is_used_for_exam:
                        messages.error(
                            request, 
                            f'Birth certificate {birth_cert_number} has already been used for registration.'
                        )
                        return redirect('candidate_create')
                    
                    # Verify name matches
                    if (birth_certificate.first_name.upper() != first_name.upper() or
                        birth_certificate.last_name.upper() != last_name.upper()):
                        messages.warning(
                            request,
                            'Birth certificate name does not match provided name. Please verify.'
                        )
                    else:
                        is_birth_cert_verified = True
                    
                except BirthCertificateRegistry.DoesNotExist:
                    messages.error(
                        request,
                        f'Birth certificate {birth_cert_number} not found in registry.'
                    )
                    return redirect('candidate_create')
            
            # Create candidate
            candidate = Candidate.objects.create(
                school=school,
                education_level=education_level,
                academic_year=academic_year,
                first_name=first_name.strip().upper(),
                middle_name=middle_name.strip().upper(),
                last_name=last_name.strip().upper(),
                gender=gender,
                date_of_birth=dob,
                phone_number=phone_number,
                parent_guardian_phone=parent_guardian_phone,
                birth_certificate=birth_certificate,
                is_birth_cert_verified=is_birth_cert_verified,
                registered_by=request.user
            )
            
            # Mark birth certificate as used
            if birth_certificate:
                birth_certificate.is_used_for_exam = True
                birth_certificate.used_exam_level = education_level
                birth_certificate.save()
            
            # Log activity
            log_activity(
                user=request.user,
                action='CREATE',
                description=f'Created candidate: {candidate.index_number} - {candidate.get_full_name()}',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            messages.success(
                request,
                f'Candidate registered successfully! Index Number: {candidate.index_number}'
            )
            return redirect('candidate_detail', index_number=candidate.index_number)
            
        except Exception as e:
            messages.error(request, f'Error creating candidate: {str(e)}')
            return redirect('candidate_create')
    
    # GET request - show form
    schools = School.objects.filter(is_active=True).order_by('name')
    education_levels = EducationLevel.objects.filter(is_active=True)
    academic_years = AcademicYear.objects.all().order_by('-year')
    
    context = {
        'schools': schools,
        'education_levels': education_levels,
        'academic_years': academic_years,
        'mode': 'create',
        'now': timezone.now(),
    }
    
    return render(request, 'candidates/candidate_form.html', context)


@login_required
def candidate_update(request, index_number):
    """Update candidate information"""
    candidate = get_object_or_404(Candidate, index_number=index_number)
    
    if request.method == 'POST':
        try:
            # Update fields
            candidate.first_name = request.POST.get('first_name', '').strip().upper()
            candidate.middle_name = request.POST.get('middle_name', '').strip().upper()
            candidate.last_name = request.POST.get('last_name', '').strip().upper()
            candidate.gender = request.POST.get('gender')
            candidate.phone_number = request.POST.get('phone_number', '')
            candidate.parent_guardian_phone = request.POST.get('parent_guardian_phone')
            
            # Update date of birth if provided
            dob_str = request.POST.get('date_of_birth')
            if dob_str:
                candidate.date_of_birth = datetime.strptime(dob_str, '%Y-%m-%d').date()
            
            # Update status if admin
            if request.user.is_admin or request.user.is_knec_staff:
                is_active = request.POST.get('is_active') == 'on'
                candidate.is_active = is_active
            
            candidate.save()
            
            # Log activity
            log_activity(
                user=request.user,
                action='UPDATE',
                description=f'Updated candidate: {candidate.index_number}',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            messages.success(request, 'Candidate updated successfully!')
            return redirect('candidate_detail', index_number=candidate.index_number)
            
        except Exception as e:
            messages.error(request, f'Error updating candidate: {str(e)}')
            return redirect('candidate_update', index_number=index_number)
    
    # GET request - show form
    schools = School.objects.filter(is_active=True).order_by('name')
    education_levels = EducationLevel.objects.filter(is_active=True)
    academic_years = AcademicYear.objects.all().order_by('-year')
    
    context = {
        'candidate': candidate,
        'schools': schools,
        'education_levels': education_levels,
        'academic_years': academic_years,
        'mode': 'update',
        'now': timezone.now(),
    }
    
    return render(request, 'candidates/candidate_form.html', context)


@login_required
def candidate_delete(request, index_number):
    """Delete candidate (AJAX only)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    # Check permissions
    if not (request.user.is_admin or request.user.is_knec_staff):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        candidate = get_object_or_404(Candidate, index_number=index_number)
        candidate_name = candidate.get_full_name()
        
        # Log activity before deletion
        log_activity(
            user=request.user,
            action='DELETE',
            description=f'Deleted candidate: {candidate.index_number} - {candidate_name}',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        # Mark birth certificate as unused if exists
        if candidate.birth_certificate:
            birth_cert = candidate.birth_certificate
            birth_cert.is_used_for_exam = False
            birth_cert.used_exam_level = None
            birth_cert.save()
        
        candidate.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Candidate {index_number} deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def export_candidates_to_excel(queryset, request):
    """Export candidates to Excel file"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"
    
    # Define header style
    header_fill = PatternFill(start_color="2c5aa0", end_color="2c5aa0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Define headers
    headers = [
        'Index Number', 'First Name', 'Middle Name', 'Last Name', 'Gender',
        'Date of Birth', 'School Code', 'School Name', 'Education Level',
        'Academic Year', 'Birth Certificate', 'Cert Verified', 'Phone Number',
        'Parent/Guardian Phone', 'Status', 'Registration Date'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row_num, candidate in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = candidate.index_number
        ws.cell(row=row_num, column=2).value = candidate.first_name
        ws.cell(row=row_num, column=3).value = candidate.middle_name
        ws.cell(row=row_num, column=4).value = candidate.last_name
        ws.cell(row=row_num, column=5).value = candidate.get_gender_display()
        ws.cell(row=row_num, column=6).value = candidate.date_of_birth.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=7).value = candidate.school.code
        ws.cell(row=row_num, column=8).value = candidate.school.name
        ws.cell(row=row_num, column=9).value = str(candidate.education_level)
        ws.cell(row=row_num, column=10).value = candidate.academic_year.year
        ws.cell(row=row_num, column=11).value = (
            candidate.birth_certificate.certificate_number 
            if candidate.birth_certificate else 'N/A'
        )
        ws.cell(row=row_num, column=12).value = 'Yes' if candidate.is_birth_cert_verified else 'No'
        ws.cell(row=row_num, column=13).value = candidate.phone_number
        ws.cell(row=row_num, column=14).value = candidate.parent_guardian_phone
        ws.cell(row=row_num, column=15).value = 'Active' if candidate.is_active else 'Inactive'
        ws.cell(row=row_num, column=16).value = candidate.registration_date.strftime('%Y-%m-%d %H:%M')
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = len(headers[col_num - 1]) + 2
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 2)
        ws.column_dimensions[column_letter].width = min(max_length, 50)
    
    # Log activity
    log_activity(
        user=request.user,
        action='EXPORT',
        description=f'Exported {queryset.count()} candidates to Excel',
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'candidates_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response